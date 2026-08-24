"""Generate a synthetic exam export at any scale, for benchmarking ingest.

    python tools/make_load.py --alerts 50000 --frames 20000
    python tools/make_load.py --alerts 200000 --frames 100000 --zip

Produces, under tools/_load/:
    alerts_<n>.xlsx          alert export, real header spellings
    evidence_<n>/            per-modality folders of frames, real filenames
    evidence_<n>.zip         the same drop as an archive (--zip)
    roster_<n>.xlsx          official centre list (the true denominator)

Shapes match what the ingest actually parses: the Alarm ID carries the channel
(C-{centreSeq}-{channel}-{epochMillis}-{seq}), frames are named
'{AlarmID}_{dssId}.jpg', and timestamps use one of the accepted spellings.
"""
from __future__ import annotations

import argparse
import random
import shutil
import zipfile
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent
OUT = ROOT / "_load"

# channel -> (folder, alarm-type label) for the shipped catalogue
MODALITIES = [
    ("2",  "ZI",  "Zone Intrusion"),
    ("14", "MD",  "Mobile Detection"),
    ("11", "CT",  "Camera Tampering"),
    ("3",  "TP",  "Trunk Placed"),
    ("4",  "TO",  "Trunk Open"),
    ("1",  "CD",  "Crowd Detection"),
    ("6",  "NP",  "No Person Detection"),
    ("12", "INM", "Invigilator Not Moving"),
]

STATES = ["Uttar Pradesh", "Bihar", "Rajasthan", "Madhya Pradesh", "Haryana"]
DISTRICTS = {
    "Uttar Pradesh": ["Lucknow", "Kanpur Nagar", "Varanasi", "Agra", "Prayagraj", "Meerut", "Gorakhpur"],
    "Bihar": ["Patna", "Gaya", "Muzaffarpur", "Bhagalpur"],
    "Rajasthan": ["Jaipur", "Jodhpur", "Kota", "Ajmer"],
    "Madhya Pradesh": ["Bhopal", "Indore", "Gwalior", "Jabalpur"],
    "Haryana": ["Gurugram", "Faridabad", "Hisar", "Panipat"],
}
ZONES = ["Exam Hall", "Corridor", "Entrance", "Strong Room", "Control Room"]

# smallest valid JPEG that Pillow will open — the content is irrelevant to ingest,
# only the filename and the count matter
JPEG = bytes.fromhex(
    "ffd8ffe000104a46494600010100000100010000ffdb004300ffffffffffffffffffffffffff"
    "ffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffff"
    "ffffffffffffffffffffffffffffffffffffffc00011080001000103012200021101031101ff"
    "c4001f0000010501010101010100000000000000000102030405060708090a0bffc400b51000"
    "02010303020403050504040000017d01020300041105122131410613516107227114328191a1"
    "082342b1c11552d1f02433627282090a161718191a25262728292a3435363738393a43444546"
    "4748494a535455565758595a636465666768696a737475767778797a838485868788898a9293"
    "9495969798999aa2a3a4a5a6a7a8a9aab2b3b4b5b6b7b8b9bac2c3c4c5c6c7c8c9cad2d3d4d5"
    "d6d7d8d9dae1e2e3e4e5e6e7e8e9eaf1f2f3f4f5f6f7f8f9faffda000801010000003f00bf80"
    "0fffd9")


def build(alerts: int, frames: int, centres: int, days: int, make_zip: bool) -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    rng = random.Random(20260823)
    start = datetime(2026, 6, 14, 7, 0, 0)

    xl = OUT / f"alerts_{alerts}.xlsx"
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet("Alerts")
    ws.append(["Alarm ID", "Center Code", "Center Name", "District/City", "State",
               "Camera Name", "Camera Sub Location", "Alert Timestamp (IST)",
               "Alarm Type", "DSS ID", "Ticket ID"])

    centre_meta = []
    for i in range(centres):
        st = rng.choice(STATES)
        centre_meta.append((f"C{i:05d}", f"Centre {i:05d}", rng.choice(DISTRICTS[st]), st))

    ids: list[tuple[str, str, str]] = []   # (alarm_id, dss_id, folder)
    for i in range(alerts):
        ch, folder, label = MODALITIES[i % len(MODALITIES)]
        cseq = rng.randrange(centres)
        code, cname, dist, st = centre_meta[cseq]
        when = start + timedelta(days=rng.randrange(days),
                                 seconds=rng.randrange(0, 11 * 3600))
        alarm_id = f"C-{cseq:05d}-{ch}-{int(when.timestamp() * 1000)}-{i:07d}"
        dss = f"D{i:08d}"
        ws.append([alarm_id, code, cname, dist, st,
                   f"CAM-{rng.randrange(1, 9):02d}", rng.choice(ZONES),
                   when.strftime("%d/%m/%Y, %H:%M:%S"), label, dss, f"T{i:07d}"])
        if len(ids) < frames:
            ids.append((alarm_id, dss, folder))
    wb.save(xl)

    ev = OUT / f"evidence_{frames}"
    if ev.exists():
        shutil.rmtree(ev)
    for _, _, folder in MODALITIES and [(0, 0, m[1]) for m in MODALITIES]:
        (ev / folder).mkdir(parents=True, exist_ok=True)
    for alarm_id, dss, folder in ids:
        (ev / folder / f"{alarm_id}_{dss}.jpg").write_bytes(JPEG)

    rl = OUT / f"roster_{centres}.xlsx"
    rwb = openpyxl.Workbook(write_only=True)
    rws = rwb.create_sheet("Centres")
    rws.append(["Centre Code", "Centre Name", "District", "State"])
    for code, cname, dist, st in centre_meta:
        rws.append([code, cname, dist, st])
    rwb.save(rl)

    zp = None
    if make_zip:
        zp = OUT / f"evidence_{frames}.zip"
        with zipfile.ZipFile(zp, "w", zipfile.ZIP_STORED) as z:
            for p in sorted(ev.rglob("*")):
                if p.is_file():
                    z.write(p, p.relative_to(ev).as_posix())

    mb = lambda p: p.stat().st_size / 1048576
    print(f"  {xl.name:<28} {alerts:>8,} rows   {mb(xl):6.1f} MB")
    print(f"  {ev.name + '/':<28} {len(ids):>8,} frames")
    print(f"  {rl.name:<28} {centres:>8,} centres  {mb(rl):6.1f} MB")
    if zp:
        print(f"  {zp.name:<28} {'':>8}          {mb(zp):6.1f} MB")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--alerts", type=int, default=50_000)
    ap.add_argument("--frames", type=int, default=20_000)
    ap.add_argument("--centres", type=int, default=1_200)
    ap.add_argument("--days", type=int, default=2)
    ap.add_argument("--zip", action="store_true")
    a = ap.parse_args()
    build(a.alerts, a.frames, a.centres, a.days, a.zip)
