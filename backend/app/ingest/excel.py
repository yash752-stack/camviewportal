"""Schema-driven Excel reader.

Exam exports do not share a column layout: UPESSC TGT has ten columns, UPPSC GIC
Lecturer has twelve with an added Camera Name and Ticket ID. Columns are therefore
resolved by header label, never by position. Timestamp wording also differs
(24-hour vs 12-hour am/pm) and is normalised to a single datetime on read. The
workbook is opened read-only and streamed so a 30k-row sheet never loads whole.
"""
from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterator

import openpyxl

# Header label -> canonical field. Each canonical field lists every spelling seen
# across exports; lookup is case- and whitespace-insensitive.
HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "centre_code": ("center code", "centre code"),
    "centre_name": ("center name", "centre name"),
    "district": ("district/city", "district", "district / city"),
    "state": ("state",),
    "camera_name": ("camera name",),
    "zone": ("camera sub location", "camera sublocation", "sub location"),
    "timestamp": ("alert timestamp (ist)", "alert timestamp", "timestamp"),
    "alarm_type": ("alarm type", "alert type"),
    "ticket_id": ("ticket id",),
    "dss_id": ("dss id",),
    "alarm_id": ("alarm id", "alert id"),
}

REQUIRED = ("centre_code", "district", "timestamp", "alarm_type", "alarm_id")

# Accepted timestamp spellings, tried in order.
_TS_FORMATS = (
    "%d/%m/%Y, %I:%M:%S %p",   # 14/6/2026, 10:00:15 am
    "%d/%m/%Y, %H:%M:%S",       # 3/6/2026, 16:29:59
    "%d/%m/%Y %I:%M:%S %p",
    "%d/%m/%Y %H:%M:%S",
)


class SchemaError(ValueError):
    """Raised when a required column cannot be located in a worksheet."""


@dataclass
class RawAlert:
    alarm_id: str
    channel: str
    centre_code: str
    centre_name: str
    district: str
    state: str
    zone: str
    camera_name: str
    alarm_type: str
    timestamp: datetime | None
    timestamp_raw: str
    dss_id: str
    ticket_id: str
    sheet: str


def _norm(value: object) -> str:
    return "" if value is None else str(value).strip()


def _build_column_map(header_row: tuple) -> dict[str, int]:
    seen = {}
    for index, cell in enumerate(header_row):
        label = _norm(cell).lower()
        if not label:
            continue
        for field, aliases in HEADER_ALIASES.items():
            if label in aliases and field not in seen:
                seen[field] = index
    missing = [f for f in REQUIRED if f not in seen]
    if missing:
        present = [_norm(c) for c in header_row if _norm(c)]
        raise SchemaError(
            f"missing required column(s): {', '.join(missing)}; "
            f"headers present: {', '.join(present)}"
        )
    return seen


def _parse_timestamp(raw: str) -> datetime | None:
    text = raw.strip()
    if not text:
        return None
    candidate = text.replace(" am", " AM").replace(" pm", " PM")
    for fmt in _TS_FORMATS:
        try:
            return datetime.strptime(candidate, fmt)
        except ValueError:
            continue
    return None


def _channel_from_alarm_id(alarm_id: str) -> str:
    # C-{centreSeq}-{channel}-{epochMillis}-{seq}
    parts = alarm_id.split("-")
    return parts[2] if len(parts) >= 5 else ""


def read_workbook(path: Path) -> Iterator[RawAlert]:
    """Yield one RawAlert per data row across every sheet in the workbook."""
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet in workbook.worksheets:
            rows = sheet.iter_rows(values_only=True)
            try:
                header = next(rows)
            except StopIteration:
                continue
            cols = _build_column_map(header)

            def cell(row: tuple, field: str) -> str:
                idx = cols.get(field)
                return _norm(row[idx]) if idx is not None and idx < len(row) else ""

            for row in rows:
                alarm_id = cell(row, "alarm_id")
                centre_code = cell(row, "centre_code")
                if not alarm_id or not centre_code:
                    continue
                ts_raw = cell(row, "timestamp")
                yield RawAlert(
                    alarm_id=alarm_id,
                    channel=_channel_from_alarm_id(alarm_id),
                    centre_code=centre_code,
                    centre_name=cell(row, "centre_name"),
                    district=cell(row, "district"),
                    state=cell(row, "state"),
                    zone=cell(row, "zone"),
                    camera_name=cell(row, "camera_name"),
                    alarm_type=cell(row, "alarm_type"),
                    timestamp=_parse_timestamp(ts_raw),
                    timestamp_raw=ts_raw,
                    dss_id=cell(row, "dss_id"),
                    ticket_id=cell(row, "ticket_id"),
                    sheet=sheet.title,
                )
    finally:
        workbook.close()


# Header spellings seen in official centre-list exports (a different file from the
# alert export). Only a centre-code column is mandatory.
_ROSTER_ALIASES: dict[str, tuple[str, ...]] = {
    "centre_code": ("center code", "centre code", "centre id", "center id", "venue code", "code"),
    "centre_name": ("center name", "centre name", "centre", "center", "name of centre",
                    "name of center", "venue", "venue name", "school name", "centre name "),
    "district": ("district/city", "district", "district / city", "city"),
    "state": ("state",),
}


def read_centre_roster(path: Path) -> list[dict]:
    """Read the authoritative centre list. Resolves columns by header label like
    read_workbook, but only a centre-code column is required; name/district/state
    are taken when present. Returns de-duplicated rows keyed by centre code."""
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out: dict[str, dict] = {}
    try:
        for sheet in workbook.worksheets:
            rows = sheet.iter_rows(values_only=True)
            try:
                header = next(rows)
            except StopIteration:
                continue
            cmap: dict[str, int] = {}
            for index, cell in enumerate(header):
                label = _norm(cell).lower()
                if not label:
                    continue
                for field, aliases in _ROSTER_ALIASES.items():
                    if label in aliases and field not in cmap:
                        cmap[field] = index
            if "centre_code" not in cmap:
                continue   # not a roster sheet — skip it

            def g(row: tuple, field: str) -> str:
                idx = cmap.get(field)
                return _norm(row[idx]) if idx is not None and idx < len(row) else ""

            for row in rows:
                code = g(row, "centre_code")
                if not code:
                    continue
                out.setdefault(code, {"centre_code": code, "centre_name": g(row, "centre_name"),
                                      "district": g(row, "district"), "state": g(row, "state")})
    finally:
        workbook.close()
    return list(out.values())
